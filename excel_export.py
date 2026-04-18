"""
Excel report generation for DataNanite insight export.

build_excel_report(results, title) → bytes (.xlsx)

Workbook layout:
  Sheet 1 : Dashboard  — title, generated time, per-result KPI summary, metric tiles
  Sheet N : [Result N] — banner, styled data table, matplotlib chart image below
  Pivot_N : cross-tab  — auto-created when data has 2 categorical + 1 numeric col

Charts are rendered as matplotlib PNG images and embedded in sheets, giving full
control over colours, annotations, axes, legends, and number formatting.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand colours ──────────────────────────────────────────────────────────────
_DARK      = "1E293B"
_WHITE     = "F8FAFC"
_ALT       = "F1F5F9"
_ACCENT    = "6366F1"
_ACCENT_BG = "EEF2FF"
_MUTE      = "64748B"
_MAX_ROWS  = 10_000

# Categorical palette (hex strings, no #)
_PALETTE = [
    "#2563EB",  # blue
    "#DC2626",  # red
    "#16A34A",  # green
    "#CA8A04",  # amber
    "#9333EA",  # purple
    "#0891B2",  # cyan
    "#EA580C",  # orange
    "#0D9488",  # teal
    "#DB2777",  # pink
    "#4338CA",  # indigo
]

# ── Sheet style helpers ────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border() -> Border:
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row, col)
        c.fill      = _fill(_DARK)
        c.font      = Font(bold=True, color=_WHITE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 20

def _col_width(col_name: str, rows: List[Dict]) -> float:
    sample = [str(r.get(col_name, "")) for r in rows[:20]]
    return min(max(len(col_name) + 2, max((len(v) for v in sample), default=0) + 2, 8), 42)

def _safe_name(text: str, idx: int) -> str:
    name = re.sub(r'[\\/*?:\[\]]', "", text).strip()[:28]
    return name or f"Result {idx + 1}"


# ── Column classification ──────────────────────────────────────────────────────

def _is_numeric(val) -> bool:
    if val is None:
        return False
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False

def _num_cols(cols: List[str], rows: List[Dict]) -> List[str]:
    sample = rows[:8]
    return [c for c in cols
            if sample and all(_is_numeric(r.get(c)) for r in sample if r.get(c) is not None)]

def _cat_cols(cols: List[str], num: List[str]) -> List[str]:
    return [c for c in cols if c not in num]

def _human_label(col: str) -> str:
    return col.replace("_", " ").title()

def _short_label(text: str, max_len: int = 18) -> str:
    """Word-aware truncation."""
    if len(text) <= max_len:
        return text
    cut = text[:max_len].rsplit(" ", 1)[0]
    if len(cut) < max_len - 6:
        cut = text[:max_len - 1]
    return cut + "…"


# ── Chart type detection ───────────────────────────────────────────────────────

def _detect_chart(cols: List[str], rows: List[Dict]) -> Optional[Dict]:
    if not rows or len(cols) < 2:
        return None

    num = _num_cols(cols, rows)
    cat = _cat_cols(cols, num)
    lbl = cat[0] if cat else None

    if lbl:
        is_id = bool(re.search(r"(_id|_key|_sk|_sku|_code|_uuid|_no|_num|_ref|_pk)$", lbl, re.I)
                     or re.fullmatch(r"id|key|pk", lbl, re.I))
        if is_id:
            return None

    if len(num) == 2 and not lbl and len(rows) >= 6:
        return {"type": "scatter", "x": num[0], "y": num[1]}

    if not lbl or not num:
        return None

    ll = lbl.lower()

    if len(rows) == 1 and len(num) >= 2:
        return None  # KPI tiles only — no chart

    if re.search(r"date|month|year|week|day|quarter|period|time|fiscal|yr|qtr|wk", ll):
        return {"type": "line", "lbl": lbl, "num": num[:5]}

    if re.search(r"bucket|bin|band|range|bracket|tier|cohort|decile|quartile|percentile", ll):
        return {"type": "vbar", "lbl": lbl, "num": num[:1]}

    all_names = " ".join(cols).lower()
    if re.search(r"variance|delta|diff|deviation|bridge|impact|contrib|change|opening|closing|movement", all_names) and len(rows) <= 25:
        return {"type": "hbar", "lbl": lbl, "num": num[:1]}

    if len(num) >= 2 and len(rows) <= 30:
        is_comp = bool(re.search(r"channel|segment|region|category|brand|product|division|dept|territory|country|market", ll))
        return {"type": "stacked" if is_comp else "grouped", "lbl": lbl, "num": num[:6]}

    is_share = bool(re.search(r"channel|segment|region|category|brand|type|division", ll))
    if is_share and len(rows) <= 8 and len(num) == 1:
        return {"type": "doughnut", "lbl": lbl, "num": num}

    if len(rows) <= 8 and len(num) == 1:
        return {"type": "vbar", "lbl": lbl, "num": num}

    if len(rows) <= 80 and len(num) == 1:
        return {"type": "hbar", "lbl": lbl, "num": num}

    return None


# ── Matplotlib chart engine ────────────────────────────────────────────────────

_FIG_W   = 11.0   # inches
_FIG_H   = 6.5
_DPI     = 130
_FGCOLOR = "#111827"
_AXCOLOR = "#374151"
_GRID    = "#E5E7EB"


def _fmt_val(v: float, col_names: List[str]) -> str:
    combined = " ".join(col_names).lower()
    if re.search(r"pct|percent|rate|ratio|share", combined):
        return f"{v:.1%}"
    if re.search(r"variance_rate|rate$", combined):
        return f"{v:.1%}"
    if abs(v) >= 1_000_000:
        return f"{v / 1_000_000:.2f}M"
    if abs(v) >= 1_000:
        return f"{v / 1_000:.1f}K"
    return f"{v:,.0f}"


def _axis_fmt(col_names: List[str]):
    combined = " ".join(col_names).lower()
    is_pct   = bool(re.search(r"pct|percent|rate|ratio|share", combined))
    def _f(x, pos):
        if is_pct:
            return f"{x:.0%}"
        if abs(x) >= 1_000_000:
            return f"{x / 1_000_000:.1f}M"
        if abs(x) >= 1_000:
            return f"{x / 1_000:.0f}K"
        return f"{x:,.0f}"
    return mticker.FuncFormatter(_f)


def _bar_colors(values: List[float], col_names: List[str]) -> List[str]:
    """Semantic: variance cols → red/green diverging. Others → intensity-scaled blue."""
    combined = " ".join(col_names).lower()
    is_var   = bool(re.search(r"variance|delta|diff|deviation|over|under|gap", combined))
    if is_var:
        return ["#DC2626" if v > 0 else "#16A34A" if v < 0 else "#9CA3AF" for v in values]
    # Intensity-scaled single hue: more intense = higher value
    if not values:
        return [_PALETTE[0]] * len(values)
    max_abs = max(abs(v) for v in values) or 1
    return [f"#{int(0x1e + 0x37 * abs(v)/max_abs):02x}"
            f"{int(0x25 + 0x40 * (1 - abs(v)/max_abs)):02x}"
            f"{int(0x63 + 0x3c * (1 - abs(v)/max_abs)):02x}"
            for v in values]


def _apply_style(ax, title: str, xlabel: str = "", ylabel: str = "", grid_axis: str = "y"):
    ax.set_title(title, fontsize=11.5, fontweight="bold", pad=14,
                 color=_FGCOLOR, loc="left")
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=9, color=_AXCOLOR, labelpad=6)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=9, color=_AXCOLOR, labelpad=6)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if grid_axis == "x":
        ax.spines["left"].set_visible(False)
        ax.grid(axis="x", color=_GRID, linewidth=0.8, zorder=0)
        ax.tick_params(axis="y", left=False, labelsize=8.5, colors=_AXCOLOR)
        ax.tick_params(axis="x", labelsize=8, colors=_AXCOLOR)
    else:
        ax.spines["bottom"].set_visible(False)
        ax.grid(axis="y", color=_GRID, linewidth=0.8, zorder=0)
        ax.tick_params(axis="x", bottom=False, labelsize=8.5, colors=_AXCOLOR)
        ax.tick_params(axis="y", labelsize=8, colors=_AXCOLOR)
    ax.set_axisbelow(True)
    ax.figure.patch.set_facecolor("white")
    ax.set_facecolor("white")


def _to_png(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=_DPI, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _chart_hbar(labels: List[str], values: List[float],
                title: str, xlabel: str, col_names: List[str]) -> bytes:
    n     = len(labels)
    h     = max(5.0, min(n * 0.55 + 1.8, 15.0))
    fig, ax = plt.subplots(figsize=(_FIG_W, h))

    colors  = _bar_colors(values, col_names)
    y_pos   = list(range(n))
    # Draw highest-value bar at top → reverse so largest is at top
    rev_lbl = list(reversed(labels))
    rev_val = list(reversed(values))
    rev_col = list(reversed(colors))

    bars = ax.barh(y_pos, rev_val, color=rev_col, height=0.62, zorder=3,
                   edgecolor="white", linewidth=0.5)

    # Value labels at end of each bar
    max_abs = max(abs(v) for v in values) if values else 1
    for bar, val in zip(bars, rev_val):
        w = bar.get_width()
        ax.text(w + max_abs * 0.012, bar.get_y() + bar.get_height() / 2,
                _fmt_val(val, col_names),
                va="center", ha="left", fontsize=8, fontweight="600", color=_FGCOLOR)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(rev_lbl, fontsize=8.5)
    ax.xaxis.set_major_formatter(_axis_fmt(col_names))
    ax.set_xlim(0, max_abs * 1.20)
    ax.tick_params(axis="y", labelsize=8.5)

    # Add a subtle reference line at 0 for variance charts
    combined = " ".join(col_names).lower()
    if re.search(r"variance|delta|diff", combined):
        ax.axvline(0, color="#6B7280", linewidth=0.8, zorder=2)

    # Legend for diverging (variance) charts
    if re.search(r"variance|delta|diff|deviation", combined):
        import matplotlib.patches as mpatches
        patches = [
            mpatches.Patch(color="#DC2626", label="Over Budget"),
            mpatches.Patch(color="#16A34A", label="Under Budget"),
        ]
        ax.legend(handles=patches, loc="lower right", fontsize=8,
                  framealpha=0.9, fancybox=True)

    _apply_style(ax, title, xlabel=xlabel, grid_axis="x")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _chart_vbar(labels: List[str], values: List[float],
                title: str, ylabel: str, col_names: List[str]) -> bytes:
    fig, ax = plt.subplots(figsize=(_FIG_W, _FIG_H))
    colors  = _bar_colors(values, col_names)
    x_pos   = np.arange(len(labels))
    bars    = ax.bar(x_pos, values, color=colors, width=0.62, zorder=3,
                     edgecolor="white", linewidth=0.5)

    max_abs = max(abs(v) for v in values) if values else 1
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max_abs * 0.015,
                _fmt_val(val, col_names),
                ha="center", va="bottom", fontsize=8, fontweight="600", color=_FGCOLOR)

    long_labels = any(len(l) > 8 for l in labels)
    ax.set_xticks(x_pos)
    ax.set_xticklabels(labels, fontsize=8.5, rotation=30 if long_labels else 0,
                       ha="right" if long_labels else "center")
    ax.yaxis.set_major_formatter(_axis_fmt(col_names))
    ax.set_ylim(0, max_abs * 1.18)
    _apply_style(ax, title, ylabel=ylabel, grid_axis="y")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _chart_line(x_labels: List[str], series: Dict[str, List[float]],
                title: str, xlabel: str, ylabel: str, col_names: List[str]) -> bytes:
    fig, ax = plt.subplots(figsize=(_FIG_W, _FIG_H))

    for i, (name, vals) in enumerate(series.items()):
        color = _PALETTE[i % len(_PALETTE)]
        ax.plot(range(len(x_labels)), vals, color=color, linewidth=2.2,
                marker="o", markersize=5.5, label=_human_label(name),
                markeredgecolor="white", markeredgewidth=0.8, zorder=4)
        if i == 0:
            ax.fill_between(range(len(x_labels)), vals, alpha=0.08, color=color)
        # Annotate last value
        if vals:
            ax.annotate(_fmt_val(vals[-1], [name]),
                        xy=(len(vals) - 1, vals[-1]),
                        xytext=(6, 3), textcoords="offset points",
                        fontsize=7.5, color=color, fontweight="600")

    long_x = any(len(l) > 6 for l in x_labels)
    ax.set_xticks(range(len(x_labels)))
    ax.set_xticklabels(x_labels, fontsize=8,
                       rotation=35 if long_x else 0,
                       ha="right" if long_x else "center")
    ax.yaxis.set_major_formatter(_axis_fmt(col_names))
    if len(series) > 1:
        ax.legend(fontsize=8.5, framealpha=0.9, loc="best",
                  fancybox=True, frameon=True)
    _apply_style(ax, title, xlabel=xlabel, ylabel=ylabel, grid_axis="y")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _chart_grouped(labels: List[str], series: Dict[str, List[float]],
                   title: str, xlabel: str, ylabel: str) -> bytes:
    n_s   = len(series)
    fig, ax = plt.subplots(figsize=(_FIG_W, _FIG_H))
    x      = np.arange(len(labels))
    w      = min(0.75 / n_s, 0.32)
    for i, (name, vals) in enumerate(series.items()):
        offset = (i - n_s / 2 + 0.5) * w
        bars = ax.bar(x + offset, vals, w, label=_human_label(name),
                      color=_PALETTE[i % len(_PALETTE)], zorder=3,
                      edgecolor="white", linewidth=0.4)
        # Value labels for short series
        if n_s <= 3:
            max_v = max(abs(v) for v in vals) if vals else 1
            for bar, val in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height() + max_v * 0.015,
                        _fmt_val(val, [name]),
                        ha="center", va="bottom", fontsize=7, color=_FGCOLOR)

    long_labels = any(len(l) > 8 for l in labels)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=8.5,
                       rotation=30 if long_labels else 0,
                       ha="right" if long_labels else "center")
    ax.yaxis.set_major_formatter(_axis_fmt(list(series.keys())))
    ax.legend(fontsize=8.5, framealpha=0.9, loc="upper right", fancybox=True)
    _apply_style(ax, title, xlabel=xlabel, ylabel=ylabel, grid_axis="y")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _chart_stacked(labels: List[str], series: Dict[str, List[float]],
                   title: str, xlabel: str, ylabel: str) -> bytes:
    fig, ax = plt.subplots(figsize=(_FIG_W, _FIG_H))
    x      = np.arange(len(labels))
    bottom = np.zeros(len(labels))
    for i, (name, vals) in enumerate(series.items()):
        arr = np.array(vals, dtype=float)
        ax.bar(x, arr, 0.65, bottom=bottom,
               label=_human_label(name),
               color=_PALETTE[i % len(_PALETTE)], zorder=3,
               edgecolor="white", linewidth=0.4)
        bottom += arr

    long_labels = any(len(l) > 8 for l in labels)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=8.5,
                       rotation=30 if long_labels else 0,
                       ha="right" if long_labels else "center")
    ax.yaxis.set_major_formatter(_axis_fmt(list(series.keys())))
    ax.legend(fontsize=8.5, framealpha=0.9, loc="upper right", fancybox=True)
    _apply_style(ax, title, xlabel=xlabel, ylabel=ylabel, grid_axis="y")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _chart_doughnut(labels: List[str], values: List[float], title: str) -> bytes:
    fig, ax = plt.subplots(figsize=(10, 6.5))
    colors  = [_PALETTE[i % len(_PALETTE)] for i in range(len(values))]
    wedges, _, autotexts = ax.pie(
        values, labels=None, autopct="%1.1f%%",
        colors=colors, startangle=90, pctdistance=0.78,
        wedgeprops=dict(width=0.52, edgecolor="white", linewidth=2.5)
    )
    for at in autotexts:
        at.set_fontsize(8.5)
        at.set_fontweight("bold")
        at.set_color("white")
    ax.legend(wedges, labels, loc="center left", bbox_to_anchor=(0.98, 0.5),
              fontsize=8.5, frameon=True, framealpha=0.9, fancybox=True)
    ax.set_title(title, fontsize=11.5, fontweight="bold", pad=14,
                 color=_FGCOLOR, loc="left")
    fig.patch.set_facecolor("white")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _chart_scatter(x_vals: List[float], y_vals: List[float],
                   x_col: str, y_col: str, title: str) -> bytes:
    fig, ax = plt.subplots(figsize=(_FIG_W, _FIG_H))
    ax.scatter(x_vals, y_vals, color=_PALETTE[0], s=55, alpha=0.72,
               zorder=3, edgecolors="white", linewidth=0.8)
    try:
        z = np.polyfit(x_vals, y_vals, 1)
        p = np.poly1d(z)
        xl = np.linspace(min(x_vals), max(x_vals), 200)
        ax.plot(xl, p(xl), color="#9CA3AF", linewidth=1.6,
                linestyle="--", zorder=2, label="Trend line")
        ax.legend(fontsize=8.5, framealpha=0.9)
    except Exception:
        pass
    ax.xaxis.set_major_formatter(_axis_fmt([x_col]))
    ax.yaxis.set_major_formatter(_axis_fmt([y_col]))
    _apply_style(ax, title,
                 xlabel=_human_label(x_col), ylabel=_human_label(y_col),
                 grid_axis="y")
    plt.tight_layout(pad=1.6)
    return _to_png(fig)


def _render_chart(cfg: Dict, rows: List[Dict], cols: List[str], title: str) -> Optional[bytes]:
    """Dispatch to the right matplotlib generator. Returns PNG bytes or None."""
    ctype = cfg["type"]
    lbl   = cfg.get("lbl")
    num   = cfg.get("num", [])
    cap   = 20 if ctype == "hbar" else 15
    rows  = rows[:cap]

    try:
        if ctype == "scatter":
            x_col, y_col = cfg["x"], cfg["y"]
            xv = [float(r.get(x_col) or 0) for r in rows]
            yv = [float(r.get(y_col) or 0) for r in rows]
            return _chart_scatter(xv, yv, x_col, y_col, title)

        if not lbl or not num:
            return None

        # Composite label when primary label has duplicates
        cat_list   = _cat_cols(cols, _num_cols(cols, rows))
        lbl_vals   = [str(r.get(lbl, "")) for r in rows]
        other_cats = [c for c in cat_list if c != lbl]
        if len(lbl_vals) != len(set(lbl_vals)) and other_cats:
            labels = [
                f"{str(r.get(lbl,'')).strip()[:6]} · {_short_label(str(r.get(other_cats[0],'')). strip(), 16)}"
                for r in rows
            ]
            if ctype in ("vbar", "grouped", "stacked"):
                ctype = "hbar"
        else:
            labels = [_short_label(str(r.get(lbl, "")), 20) for r in rows]

        xlabel = _human_label(lbl)
        ylabel = " / ".join(_human_label(c) for c in num[:2])

        if ctype == "doughnut":
            vals = [float(r.get(num[0]) or 0) for r in rows]
            return _chart_doughnut(labels, vals, title)

        if ctype == "line":
            series = {c: [float(r.get(c) or 0) for r in rows] for c in num}
            return _chart_line(labels, series, title, xlabel, ylabel, num)

        if ctype == "hbar":
            vals = [float(r.get(num[0]) or 0) for r in rows]
            return _chart_hbar(labels, vals, title, xlabel, num)

        if ctype == "vbar":
            vals = [float(r.get(num[0]) or 0) for r in rows]
            return _chart_vbar(labels, vals, title, ylabel, num)

        if ctype == "grouped":
            series = {c: [float(r.get(c) or 0) for r in rows] for c in num}
            return _chart_grouped(labels, series, title, xlabel, ylabel)

        if ctype == "stacked":
            series = {c: [float(r.get(c) or 0) for r in rows] for c in num}
            return _chart_stacked(labels, series, title, xlabel, ylabel)

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Chart render failed: %s", e)
    return None


def _embed_chart(ws, png_bytes: bytes, anchor: str):
    """Embed a PNG image into the worksheet at the given cell anchor."""
    img = XlImage(io.BytesIO(png_bytes))
    img.anchor = anchor
    ws.add_image(img)


# ── Pivot sheet ────────────────────────────────────────────────────────────────

def _maybe_pivot(wb: Workbook, cols: List[str], rows: List[Dict], base_name: str) -> bool:
    num = _num_cols(cols, rows)
    cat = _cat_cols(cols, num)
    if len(cat) != 2 or len(num) != 1:
        return False

    row_dim, col_dim, measure = cat[0], cat[1], num[0]
    col_vals = sorted({str(r.get(col_dim, "")) for r in rows})
    if len(col_vals) > 15:
        return False
    row_vals = sorted({str(r.get(row_dim, "")) for r in rows})

    agg: Dict[Tuple, float] = {}
    for r in rows:
        key = (str(r.get(row_dim, "")), str(r.get(col_dim, "")))
        try:
            agg[key] = agg.get(key, 0.0) + float(r.get(measure) or 0)
        except (TypeError, ValueError):
            pass

    ws        = wb.create_sheet(_safe_name(f"Pivot_{base_name}", 0))
    total_col = len(col_vals) + 2

    ws.cell(1, 1, row_dim)
    _style_header_row(ws, 1, total_col)
    for j, cv in enumerate(col_vals, 2):
        ws.cell(1, j, cv).alignment = Alignment(horizontal="center")
    ws.cell(1, total_col, "Total")

    for i, rv in enumerate(row_vals, 2):
        alt  = (i % 2 == 0)
        fill = _fill(_ALT) if alt else PatternFill()
        c    = ws.cell(i, 1, rv)
        c.fill, c.border, c.font = fill, _border(), Font(bold=True)

        row_total = 0.0
        for j, cv in enumerate(col_vals, 2):
            val = agg.get((rv, cv), 0.0)
            row_total += val
            cell            = ws.cell(i, j, round(val, 2))
            cell.fill       = fill
            cell.border     = _border()
            cell.alignment  = Alignment(horizontal="right")
        tot             = ws.cell(i, total_col, round(row_total, 2))
        tot.fill        = fill
        tot.border      = _border()
        tot.font        = Font(bold=True)
        tot.alignment   = Alignment(horizontal="right")

    gt = len(row_vals) + 2
    ws.cell(gt, 1, "Grand Total")
    _style_header_row(ws, gt, total_col)
    grand = 0.0
    for j, cv in enumerate(col_vals, 2):
        col_total = sum(agg.get((rv, cv), 0.0) for rv in row_vals)
        grand += col_total
        c             = ws.cell(gt, j, round(col_total, 2))
        c.fill        = _fill(_DARK)
        c.font        = Font(bold=True, color=_WHITE)
        c.border      = _border()
        c.alignment   = Alignment(horizontal="right")
    tc            = ws.cell(gt, total_col, round(grand, 2))
    tc.fill       = _fill(_DARK)
    tc.font       = Font(bold=True, color=_WHITE)
    tc.border     = _border()
    tc.alignment  = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = max(
        len(row_dim) + 2, max((len(v) for v in row_vals), default=0) + 2, 14)
    for j, cv in enumerate(col_vals, 2):
        ws.column_dimensions[get_column_letter(j)].width = max(len(cv) + 2, 10)
    ws.column_dimensions[get_column_letter(total_col)].width = 12
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    return True


# ── Data sheet ─────────────────────────────────────────────────────────────────

def _write_data_sheet(wb: Workbook, result: Dict, idx: int) -> Dict:
    desc = result.get("description") or f"Query {idx + 1}"
    cols = result.get("columns") or []
    rows = (result.get("rows") or [])[:_MAX_ROWS]
    if not cols or not rows:
        return {"name": desc, "rows": 0, "cols": 0, "kpis": []}

    name = _safe_name(desc, idx)
    ws   = wb.create_sheet(name)
    nc   = len(cols)
    nums = set(_num_cols(cols, rows))

    # Sort by primary numeric col descending (ranking) unless time-series
    num_list  = list(nums)
    cat_list  = [c for c in cols if c not in nums]
    lbl_guess = cat_list[0] if cat_list else ""
    is_time   = bool(re.search(r"date|month|year|week|day|quarter|period|time|fiscal|yr|qtr|wk",
                               lbl_guess.lower()))
    if num_list and not is_time:
        try:
            rows = sorted(rows, key=lambda r: float(r.get(num_list[0]) or 0), reverse=True)
        except (TypeError, ValueError):
            pass

    nr = len(rows)

    # Banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    b            = ws.cell(1, 1, desc)
    b.fill       = _fill(_ACCENT)
    b.font       = Font(bold=True, color=_WHITE, size=12)
    b.alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header
    for j, col in enumerate(cols, 1):
        ws.cell(2, j, col)
    _style_header_row(ws, 2, nc)

    # Data rows
    for i, row in enumerate(rows, 3):
        alt  = (i % 2 == 0)
        fill = _fill(_ALT) if alt else PatternFill()
        for j, col in enumerate(cols, 1):
            val = row.get(col)
            if val is not None and col in nums:
                try:
                    fv  = float(val)
                    val = int(fv) if fv == int(fv) else fv
                except (TypeError, ValueError):
                    pass
            c           = ws.cell(i, j, val)
            c.fill      = fill
            c.border    = _border()
            c.alignment = Alignment(
                horizontal="right" if col in nums else "left",
                vertical="center"
            )

    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = _col_width(col, rows)
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # Chart — render with matplotlib and embed as image
    cfg = _detect_chart(cols, rows)
    if cfg:
        png = _render_chart(cfg, rows, cols, desc)
        if png:
            _embed_chart(ws, png, f"A{nr + 5}")

    _maybe_pivot(wb, cols, rows, name)

    # KPI summary for dashboard
    kpis = []
    for nc_name in list(nums)[:4]:
        try:
            vals = [float(r.get(nc_name) or 0) for r in rows]
            kpis.append({
                "col": nc_name,
                "sum": round(sum(vals), 2),
                "avg": round(sum(vals) / len(vals), 2) if vals else 0,
            })
        except (TypeError, ValueError):
            pass

    return {"name": name, "desc": desc, "rows": nr, "cols": len(cols), "kpis": kpis}


# ── Dashboard sheet ────────────────────────────────────────────────────────────

def _write_dashboard(wb: Workbook, meta_list: List[Dict], title: str):
    ws = wb.create_sheet("Dashboard", 0)

    ws.merge_cells("A1:H1")
    t            = ws.cell(1, 1, title)
    t.fill       = _fill(_DARK)
    t.font       = Font(bold=True, color=_WHITE, size=16)
    t.alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    s            = ws.cell(2, 1,
                           f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}   ·   "
                           f"{len(meta_list)} result set(s)")
    s.fill       = _fill("F8FAFC")
    s.font       = Font(italic=True, color=_MUTE, size=10)
    s.alignment  = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    hdrs = ["#", "Result", "Rows", "Columns", "Top Metric", "Sum", "Avg", "Sheet"]
    for j, h in enumerate(hdrs, 1):
        ws.cell(4, j, h)
    _style_header_row(ws, 4, len(hdrs))

    for i, m in enumerate(meta_list, 5):
        alt  = (i % 2 == 0)
        fill = _fill(_ALT) if alt else PatternFill()
        vals = [
            i - 4,
            m.get("desc", m.get("name", "")),
            m.get("rows", 0),
            m.get("cols", 0),
            m["kpis"][0]["col"] if m.get("kpis") else "—",
            m["kpis"][0]["sum"] if m.get("kpis") else "—",
            m["kpis"][0]["avg"] if m.get("kpis") else "—",
            m.get("name", ""),
        ]
        for j, v in enumerate(vals, 1):
            c           = ws.cell(i, j, v)
            c.fill      = fill
            c.border    = _border()
            c.alignment = Alignment(
                horizontal="right" if j in {3, 4, 6, 7} else "left",
                vertical="center"
            )
        link            = ws.cell(i, 8)
        link.hyperlink  = f"#'{m['name']}'!A1"
        link.font       = Font(color=_ACCENT, underline="single")

    kpi_hdr = 6 + len(meta_list)
    ws.merge_cells(start_row=kpi_hdr, start_column=1, end_row=kpi_hdr, end_column=8)
    kh           = ws.cell(kpi_hdr, 1, "KEY METRICS")
    kh.fill      = _fill(_DARK)
    kh.font      = Font(bold=True, color=_WHITE, size=10)
    kh.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[kpi_hdr].height = 18

    kpi_row  = kpi_hdr + 1
    tile_col = 1
    for m in meta_list:
        for kpi in m.get("kpis", [])[:2]:
            if tile_col > 8:
                kpi_row  += 4
                tile_col  = 1
            lbl            = ws.cell(kpi_row, tile_col, kpi["col"])
            lbl.font       = Font(bold=True, size=9, color=_MUTE)
            lbl.alignment  = Alignment(horizontal="center")
            val            = ws.cell(kpi_row + 1, tile_col, kpi["sum"])
            val.font       = Font(bold=True, size=14, color=_ACCENT)
            val.fill       = _fill(_ACCENT_BG)
            val.border     = _border()
            val.alignment  = Alignment(horizontal="center")
            tile_col += 1

    for col_letter, width in zip("ABCDEFGH", [5, 42, 10, 10, 24, 14, 14, 24]):
        if (ws.column_dimensions[col_letter].width or 0) < width:
            ws.column_dimensions[col_letter].width = width
    ws.sheet_view.showGridLines = False


# ── Row normalisation ──────────────────────────────────────────────────────────

def _normalise_result(result: Dict[str, Any]) -> Dict[str, Any]:
    """Convert list-style rows [[v1,v2,...]] → dict rows [{col: v,...}]."""
    cols = result.get("columns") or []
    rows = result.get("rows") or []
    if rows and isinstance(rows[0], (list, tuple)):
        rows = [dict(zip(cols, row)) for row in rows]
    return {**result, "columns": cols, "rows": rows}


# ── Public entry point ─────────────────────────────────────────────────────────

def build_excel_report(results: List[Dict[str, Any]],
                       title: str = "DataNanite Insight Report") -> bytes:
    """
    Build a multi-tab Excel workbook from query result dicts.
    Each result: {description, columns: [str], rows: [dict | list], ...}
    Returns raw .xlsx bytes.
    """
    wb   = Workbook()
    stub = wb.active
    if stub is not None:
        wb.remove(stub)

    valid     = [_normalise_result(r) for r in results if r.get("columns") and r.get("rows")]
    meta_list = [_write_data_sheet(wb, r, i) for i, r in enumerate(valid)]

    if meta_list:
        _write_dashboard(wb, meta_list, title)
        wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
